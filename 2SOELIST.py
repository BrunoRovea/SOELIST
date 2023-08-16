#%%
# Importa bibliotecas
import pandas as pd
import glob
from collections import Counter
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk


# Cria a janela sem loop 
root = tk.Tk()


class Funcs():
    def obter_caminhoBD(self):
        # Função utilizada para obter o caminho dos arquivos Eventos.xlsx e BD_SCADA.xlsx
        # Caminho = filedialog.askopenfilename(filetypes=[('Arquivo Excel', '*.xlsx')])
        self.caminho_BD = filedialog.askopenfilename(filetypes=[('Todos os arquivos', '*.*')])
        self.loc_BT_entry.insert(tk.END, self.caminho_BD)

    def obter_caminhoEVT(self):
        # Função utilizada para obter o caminho dos arquivos Eventos.xlsx e BD_SCADA.xlsx
        # Caminho = filedialog.askopenfilename(filetypes=[('Arquivo Excel', '*.xlsx')])
        self.caminho_EVT = filedialog.askopenfilename(filetypes=[('Todos os arquivos', '*.*')])
        self.loc_EVT_entry.insert(tk.END, self.caminho_EVT)

    def carrega_bd(self): 
        # Algoritmo que importa a tagname completa e o acrônimo para Open e Close (0 e 1) do BD_SCADA

        # Seta a pasta do BD
        arquivos = glob.glob(self.caminho_BD)
        # Sinaliza o início da criação do df, isso pode demorar de 1 a 3 min
        print('---------------------------------------------------------------------------------')
        print('-----------------Criação da o dataframe SOSTAT-----------------')
        print('---------------------------------------------------------------------------------')
        # Subnam, PntNam, Acronimo
        sostat = pd.read_excel(arquivos[0], sheet_name='RANGER_SOSTAT',usecols=[2,3,18])

        # SAT1NO, STEXT0, STEXT1
        sosat = pd.read_excel(arquivos[0], sheet_name='RANGER_SOSAT1',usecols=[0,11,12])

        # Agrupa o acronimo 
        sostat_sosat = pd.merge(sostat, sosat, left_on=['ACRONM'], right_on=['SAT1NO'],how='outer')
        sostat_sosat['Tagname'] = sostat_sosat['SUBNAM']+'.'+sostat_sosat['PNTNAM']

        # Cria a variável com os dados do BD do SCADA desejados
        sostat_sosat = sostat_sosat.filter(items=['Tagname', 'STEXT0', 'STEXT1'])
        
        # Mais fácil entender este nome
        self.sostat = sostat_sosat

        # Sinaliza que a criação do df foi concluída
        print('---------------------------------------------------------------------------------')
        print('-----------------Criação da SOSTAT concluída-----------------')
        print('---------------------------------------------------------------------------------')
        
        del sostat_sosat, arquivos

    def cria_soelist(self):
        # Cria um df contendo todas as correspondências do scratch no BD_SCADA
        # Este df também possui uma coluna Event Flag
        #     Que consiste em destacar quantas correspondências um alarme 
        #     do scratch possui no BD_SCADA


        # importa o scratchpad
        scratch = pd.read_csv(self.caminho_EVT, index_col=False)
        # Extrai a coluna Event do scratchpad importado
        alarme = pd.Series(index=scratch.index, dtype=str)


        # Deixa o alarme no formato padrão da SOELIST, porém ainda truncado
        for index, row in scratch.iterrows():
            
            # Caso o ponto seja do tipo STATUS PT
            if 'STATUS PT.' in row['Event']:
                # 12 primeiros caracteres
                alarme[index] = row['Event'][12:]

                # Separa em subNam e pointNam, por 8 caracteres
                subNam   = alarme[index][:8]
                pointNam = alarme[index][8:]

                # Strip para retirar espaços em branco
                subNam = str.strip(subNam)
                pointNam = str.strip(pointNam)
                
                # Formato padrão da SOELIST subnam.pointnam (ainda truncado!)
                alarme[index] = subNam + '.' + pointNam
            else:
                # Caso seja comentário, ANALOG PT> ou outra coisa
                # Ñ há a necessidade de colocar no formato da SOELIST
                # Pois este ponto não existe no df soelist
                alarme[index] = row['Event']


        # Deleta variáveis auxiliares
        del index, row


        # Inicializa dois dicionários

        # Corresp representa todas as correspondências do alarme do DTS no BD SCADA
        corresp = dict()

        # Status é o STEXE0 e STEXT1 para cada correspondência
        status = dict()

        # Inicializa o dataframe final com o cabeçalho padrão da SOELIST
        resultado = pd.DataFrame(columns=["Event Flag", "Event Time", "Previous Event", "Status", "Tagname"])


        # Percorre a Serie alarme
        for index, i in enumerate(alarme):
            # Valores True para as correspondências do Tagname do scratchpad
            # Caso não haja correspondência, ou o valor ñ seja uma string, preenche com False
            aux = self.sostat['Tagname'].map(lambda x: i in x if isinstance(x,str) else False)

            # Caso tenha encontrado algum evento
            if aux.any():
                # Colocar a SOSTAT encontrada no formato da SOELIST
                # Divide o Event em subnam.pointnam        
                subNam   = self.sostat['Tagname'][aux].map(lambda x: x.split(".", maxsplit=1)[0] if isinstance(x, str) else "")
                pointNam = self.sostat['Tagname'][aux].map(lambda x: x.split(".", maxsplit=1)[1] if isinstance(x, str) else "")

                # strip em possíveis caracteres espúrios
                pointNam = pointNam.str.strip() 
                subNam   = subNam.str.strip() 

                # Sendo subnam necessariamente com 8 caracteres, preenchidos com espaços em branco
                subNam   = subNam + ' '*(8 - len(subNam.values[0]))

                # Preenche o dicionário corresp com todas as correspondências do ponto no formato da SOELIST
                corresp = subNam + '.' + pointNam

                # CLOSE(1) ou OPEN(0)
                # Preenche o status com o texto correspondente
                if "CLOSED" in scratch.iloc[index]['Description']:
                    status = list(self.sostat['STEXT1'][aux])

                else:
                    status = list(self.sostat['STEXT0'][aux])

                # Start time vindo do DTS
                startTime = pd.Series(scratch.iloc[index]['Start Time'])

                # Formato do SOELIST
                startTime = datetime.strptime(startTime[0], "%d/%b/%Y %H:%M:%S")
                startTime = startTime.strftime("%d/%m/%y %H:%M:%S") + '.000'

                # Índice(s) do scratch da(s) correspondência(s) na SOSTAT 
                flag = [index]*len(corresp)

                # Salva no DF final todas as correspondências com seus respectivos status e start time
                aux_df = pd.DataFrame({"Event Flag": flag, "Event Time": startTime, "Status": status, "Tagname": corresp})
                resultado = pd.concat([resultado, aux_df], ignore_index=True)

            # Caso encontre nenhuma correspondência
            else:
                # Start time vindo do DTS
                startTime = pd.Series(scratch.iloc[index]['Start Time'])

                # Formato do SOELIST
                startTime = datetime.strptime(startTime[0], "%d/%b/%Y %H:%M:%S")
                startTime = startTime.strftime("%d/%m/%y %H:%M:%S") + '.000'

                # Detecta se isso se trata de um comentário (flag = -1)
                # Ou de um ponto que deverá ser inserido pelo usuário (flag = -2)
                if ' '*48 in scratch['Description'].iloc[index]:
                    # Preenche o status com vazio
                    status = pd.Series('')
                    flag = -1
                else:
                    # Preenche o status com set ou setoverride xx.xx
                    status = scratch['Description'].iloc[index]
                    flag = -2

                # O dicionário corresp é preenchido com o comentário literal do Event
                corresp   = pd.Series(scratch.iloc[index]['Event'])

                # Preenche o dataframe com o comentário encontrado
                aux_df = pd.DataFrame({"Event Flag": flag, "Event Time": startTime, "Status": status, "Tagname": corresp})
                resultado = pd.concat([resultado, aux_df], ignore_index=True)


        # Deleta variáveis auxiliares
        # del corresp, status, aux_df, startTime, i, index, flag, aux
        # del alarme, scratch


        self.resultado_df = resultado

    def cria_tabela(self):
        # Algoritmo que plota um df em um arquivo xlsx
        # Event Flag
        #     +n para as correspondências dos scratchpads na soelist
        #     -1 para comentários
        #     -2 para eventos não encontrados no BD_SCADA (ANALOG PT. e.g.)

        # Este xlsx é separado por cores, para facilitar o usuário a preencher a soelist corretamente
        #     AZUL para alarmes com correspondência única
        #     VERDE claro/escuro para alarmes com mais de uma correspondência
        #     LARANJA para comentários
        #     VERMELHO para alarmes sem correspondência no sostat


        # Criar o arquivo XLSX de saída
        output_filename = self.getSoelist_entry.get() + '.xlsx'
        self.resultado_df.to_excel(output_filename, index=False, sheet_name='Sheet1')


        # Carregar o arquivo XLSX com openpyxl
        workbook = load_workbook(output_filename)
        worksheet = workbook['Sheet1']


        # Definir as cores de preenchimento
        orang_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='lightUp')
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        dark_green_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        blue_fill = PatternFill(start_color='91BCE3', end_color='91BCE3', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')


        # Pintar as linhas de acordo com as condições
        event_flag_counter = Counter(self.resultado_df['Event Flag'])

        for row_idx, event_flag in enumerate(self.resultado_df['Event Flag'], start=2):

            if event_flag == -1:
                for cell in worksheet[row_idx]:
                    cell.fill = orang_fill

            elif event_flag == -2:
                for cell in worksheet[row_idx]:
                    cell.fill = red_fill

            elif event_flag_counter[event_flag] > 1:
                if event_flag % 2 == 0:
                    for cell in worksheet[row_idx]:
                        cell.fill = green_fill

                else:
                    for cell in worksheet[row_idx]:
                        cell.fill = dark_green_fill

            else:
                for cell in worksheet[row_idx]:
                    cell.fill = blue_fill

        # Salva o xlsx colorido com o nome de LISTA.xlsx
        workbook.save(output_filename)


class Application(Funcs):
    def __init__(self):
        self.root = root
        self.tela()
        self.frames()
        self.widgets()
        # self.lista_frame2()

        # Coloca a janela root em loop para que a interação seja possível
        root.mainloop()
    def tela(self):
        # Cria um título para a janela
        self.root.title("Criação de scratchpads utilizando a ferramenta SOELIST")
        
        # Define a cor de fundo para a janela
        # Cores tkinter que ele entende (ggl)
        self.root.configure(background= '#1e3743')
        
        # Geometria da janela responsiva
        self.root.geometry("350x500")
        self.root.resizable(True, True)
        
        # Máximo e mínimo de responsividade
        self.root.maxsize(width= 900, height=350)
        self.root.minsize(width=500, height=200)

    def frames(self):
        self.frame_1 = tk.Frame(self.root, bd = 4, bg = '#dfe3ee', highlightbackground="#759fe6", highlightthickness=3)
        self.frame_1.place(relx = 0.02, rely = 0.02, relwidth = 0.96, relheight = 0.96)


    def widgets(self):
        self.bt_locBD = tk.Button(self.frame_1, text="Localizar BD_SCADA", bd=2, bg='#107db2', fg='white', font=('verdana', 8, 'bold'), command=self.obter_caminhoBD)
        self.bt_locBD.place(relx=0.0, rely = 0.08, relwidth = 0.3, relheight = 0.07)

        self.loc_BT_entry = tk.Entry(self.frame_1)
        self.loc_BT_entry.place(relx = 0.33, rely = 0.08, relwidth=0.65, relheight=0.07)

        self.lb_nome = tk.Label(self.frame_1, text = "Local do BD_SCADA", bg = '#dfe3ee', fg='#107db2')
        self.lb_nome.place(relx = 0.53, rely = 0.0)


        self.bt_carrBD = tk.Button(self.frame_1, text="Carregar BD_SCADA", bd=2, bg='#107db2', fg='white', font=('verdana', 8, 'bold'), command=self.carrega_bd)
        self.bt_carrBD.place(relx=0.15, rely = 0.18, relwidth = 0.65, relheight = 0.09)


        self.bt_locEVT = tk.Button(self.frame_1, text="Localizar Eventos", bd=2, bg='#107db2', fg='white', font=('verdana', 8, 'bold'), command=self.obter_caminhoEVT)
        self.bt_locEVT.place(relx=0.0, rely = 0.38, relwidth = 0.3, relheight = 0.07)

        self.loc_EVT_entry = tk.Entry(self.frame_1)
        self.loc_EVT_entry.place(relx = 0.33, rely = 0.38, relwidth=0.65, relheight=0.07)

        self.lb_nome = tk.Label(self.frame_1, text = "Local dos Eventos", bg = '#dfe3ee', fg='#107db2')
        self.lb_nome.place(relx = 0.53, rely = 0.30)


        self.bt_carrEVT = tk.Button(self.frame_1, text="Carregar Eventos", bd=2, bg='#107db2', fg='white', font=('verdana', 8, 'bold'), command=self.cria_soelist)
        self.bt_carrEVT.place(relx=0.15, rely = 0.48, relwidth = 0.65, relheight = 0.09)


        self.getSoelist_entry = tk.Entry(self.frame_1)
        self.getSoelist_entry.place(relx = 0.025, rely = 0.78, relwidth=0.95, relheight=0.07)

        self.lb_nome = tk.Label(self.frame_1, text = "Nome da lista de Eventos", bg = '#dfe3ee', fg='#107db2')
        self.lb_nome.place(relx = 0.35, rely = 0.7)


        self.getSoelist = tk.Button(self.frame_1, text="Gerar SOELIST", bd=2, bg='#107db2', fg='white', font=('verdana', 8, 'bold'), command=self.cria_tabela)
        self.getSoelist.place(relx=0.25, rely = 0.88, relwidth = 0.50, relheight = 0.1)


Application()
# %%
